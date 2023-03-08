import blJobs
import os, sys
import openpyxl

def repNormZad(id):
    err, data = blJobs.getJobListRow(id)
    repFileName = ''
    if not err:
        wb = openpyxl.load_workbook('template\\normZadTempl.xlsx')
        ws = wb['normZad']
        ws['A2'].value = f'Нормированное задание на {data.Date} г.'
        ws['A4'].value = data.Department
        ws['B11'].value = data.FIO
        ws['C11'].value = data.TabNum
        ws['D11'].value = data.Position
        ws['E11'].value = data.Level
        ws['F11'].value = data.TimeJob
        ws['I11'].value = data.TimeJob

        repFileName = f'{data.TabNum}_{data.FIO}_{data.Date}.xlsx'
        repFileName = os.path.join('reports', repFileName)
        err = ''
        if os.path.exists(repFileName):
            err = deleteFile(repFileName) 
        if not err:
            wb.save(repFileName)
      
    return err, repFileName

def openExcelFile(filename):
    err = ''
    if os.path.exists(filename):
        os.system(f'start /MIN "excel" "{filename}"')
    else:
        err = f'Нет файла: {filename}'


    return err

def printExcelFile(filename):
    os.startfile(filename, 'print')


def deleteFile(filename):
    err = ''
    try:
        os.remove(filename)
    except OSError:
        err = f"Ошибка удаления файла: {filename}"
    return err

if __name__ == '__main__':
    err, repFileName = repNormZad(1)
    if not err:
        openExcelFile(repFileName)
        # printExcelFile(repFileName)
    else:
        print (err)


